import openpyxl
import csv
import re

XLSX = '/home/diego/sources/BluenetERP/INVENTARIO GENERAL MAYO...xlsx'

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

items = []
seen_codes = set()

CI_CODE = 1
CI_NAME = 2
CI_DESC1 = 3
CI_DESC2 = 4
CI_PRECIO_IVA = 5
CI_PRECIO = 6
CI_COSTO = 7
CI_STOCK = 8
CI_STOCK2 = 9

# ── helpers ──────────────────────────────────────────────────────

def guess_uom(text):
    t = (text or '').lower()
    if re.search(r'\byarda\b', t) or re.search(r'\byard\b', t):
        return 'Yard'
    if re.search(r'\bmetro\b', t) or 'mts' in t:
        return 'Mtr'
    return 'Nos'


def clean(val):
    if val is None:
        return ''
    return str(val).strip().replace('\n', ' ').replace('\r', ' ').replace('\u2028', ' ')


def is_numeric(val):
    return val is not None and isinstance(val, (int, float))


def looks_like_code(val):
    if val is None:
        return False
    return bool(re.match(r'^[A-Za-z0-9][A-Za-z0-9_.\-/]{0,30}$', str(val).strip()))


def unique_code(base):
    c = str(base).strip()
    if c not in seen_codes:
        seen_codes.add(c)
        return c
    suffix = 1
    while f'{c}-{suffix}' in seen_codes:
        suffix += 1
    c = f'{c}-{suffix}'
    seen_codes.add(c)
    return c


def extract_item(code, name, group, description, rate, stock, uom=None):
    try:
        sv = float(stock) if stock is not None else 0
    except (ValueError, TypeError):
        return None
    if sv <= 0:
        return None

    code = unique_code(code) if code else ''
    name = clean(name) if name else code
    if not code:
        return None

    if uom is None:
        uom = guess_uom(name + ' ' + (description or ''))

    return {
        'item_code': code,
        'item_name': name[:140],
        'item_group': group,
        'stock_uom': uom,
        'description': clean(description) if description else '',
        'standard_rate': round(float(rate), 2) if rate and is_numeric(rate) else '',
        'opening_qty': int(sv),
        'warehouse': 'Guatemala',
    }


def add_item(code, name, group, description, rate, stock, uom=None):
    rec = extract_item(code, name, group, description, rate, stock, uom)
    if rec:
        items.append(rec)


# ════════════════════════════════════════════════════════════════
# 1. PLIEGOS DE PVC  (rows 6-11)
# ════════════════════════════════════════════════════════════════
for row in ws.iter_rows(min_row=6, max_row=11, values_only=True):
    vals = list(row)
    add_item(vals[CI_CODE], vals[CI_NAME], 'PLIEGOS DE PVC', vals[CI_DESC1],
             vals[CI_PRECIO_IVA], vals[CI_STOCK])

# ════════════════════════════════════════════════════════════════
# 2. EQUIPOS DE SUBLIMACION  (rows 174-180)
# ════════════════════════════════════════════════════════════════
for row in ws.iter_rows(min_row=174, max_row=180, values_only=True):
    vals = list(row)
    desc = vals[CI_DESC2] if len(vals) > CI_DESC2 and vals[CI_DESC2] else vals[CI_DESC1]
    add_item(vals[CI_CODE], vals[CI_NAME], 'EQUIPOS DE SUBLIMACIÓN', desc,
             vals[CI_PRECIO_IVA], vals[CI_STOCK])

# ════════════════════════════════════════════════════════════════
# 3. SUMINISTROS PARA SUBLIMAR  (rows 184-214)
# ════════════════════════════════════════════════════════════════
for row in ws.iter_rows(min_row=184, max_row=214, values_only=True):
    vals = list(row)
    if not vals or vals[CI_CODE] is None:
        continue

    code = vals[CI_CODE]
    if not looks_like_code(code):
        continue

    name = vals[CI_NAME] if len(vals) > CI_NAME and vals[CI_NAME] else code
    desc = vals[CI_DESC1] if len(vals) > CI_DESC1 and vals[CI_DESC1] else name
    price = vals[CI_PRECIO_IVA] if len(vals) > CI_PRECIO_IVA and is_numeric(vals[CI_PRECIO_IVA]) else None
    stock = vals[CI_STOCK] if len(vals) > CI_STOCK and is_numeric(vals[CI_STOCK]) else None

    if stock is None:
        for i in range(len(vals) - 1, 4, -1):
            if is_numeric(vals[i]):
                stock = vals[i]
                break

    add_item(code, name, 'SUMINISTROS PARA SUBLIMAR', desc, price, stock)

# ════════════════════════════════════════════════════════════════
# 4. EQUIPOS Y MAQUINARIAS — rows 329-341
# ════════════════════════════════════════════════════════════════
for row in ws.iter_rows(min_row=329, max_row=341, values_only=True):
    vals = list(row)
    if not vals or vals[1] is None:
        continue

    raw_code = vals[1]
    name = vals[2] if len(vals) > 2 and vals[2] else raw_code

    raw_code_str = clean(raw_code)
    if raw_code_str.startswith('HS CODE'):
        short = re.sub(r'[^A-Za-z0-9]', '', name).upper()
        short = re.sub(r'^ROLLO(DE)?', '', short)
        short = short[:15]
        code = short or 'MACH'
    else:
        code = raw_code

    desc_parts = []
    for idx in [3, 4]:
        if len(vals) > idx and vals[idx] and clean(vals[idx]) != clean(name):
            desc_parts.append(clean(vals[idx]))
    desc = ' '.join(desc_parts) if desc_parts else name

    price = vals[CI_PRECIO_IVA] if len(vals) > CI_PRECIO_IVA and is_numeric(vals[CI_PRECIO_IVA]) else None
    stock = vals[CI_STOCK] if len(vals) > CI_STOCK and is_numeric(vals[CI_STOCK]) else None

    if stock is None:
        for i in range(len(vals) - 1, 4, -1):
            if is_numeric(vals[i]):
                stock = vals[i]
                break

    uom = guess_uom(name + ' ' + desc)
    add_item(code, name, 'EQUIPOS Y MAQUINARIAS', desc, price, stock, uom)

# ════════════════════════════════════════════════════════════════
# 5. REPUESTOS PLOTTERS  (rows 346-407)
# ════════════════════════════════════════════════════════════════
for row in ws.iter_rows(min_row=346, max_row=407, values_only=True):
    vals = list(row)
    if not vals or vals[1] is None:
        continue

    code = vals[1]
    if not looks_like_code(code):
        continue

    name = vals[2] if len(vals) > 2 and vals[2] else code

    desc_parts = []
    for idx in [3, 4]:
        if len(vals) > idx and vals[idx] and clean(vals[idx]) != clean(name):
            desc_parts.append(clean(vals[idx]))
    desc = ' '.join(desc_parts) if desc_parts else name

    price = vals[CI_PRECIO_IVA] if len(vals) > CI_PRECIO_IVA and is_numeric(vals[CI_PRECIO_IVA]) else None

    stock = None
    for i in range(len(vals) - 1, 5, -1):
        if is_numeric(vals[i]):
            stock = int(vals[i])
            break

    if stock is None or stock <= 0:
        continue

    uom = guess_uom(name + ' ' + desc)
    add_item(code, name, 'REPUESTOS', desc, price, stock, uom)


# ════════════════════════════════════════════════════════════════
# OUTPUT FILES
# ════════════════════════════════════════════════════════════════

# ── 1. Item Groups ───────────────────────────────────────────── ─────────────────────────────────────────────
groups_seen = set()
for it in items:
    groups_seen.add(it['item_group'])

groups_csv = '/home/diego/sources/BluenetERP/erpnext_item_groups_import.csv'
with open(groups_csv, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=['item_group_name', 'parent_item_group'])
    w.writeheader()
    for g in sorted(groups_seen):
        w.writerow({'item_group_name': g, 'parent_item_group': 'All Item Groups'})
print(f'{groups_csv}  —  {len(groups_seen)} item groups')

# ── 2. Items (no opening stock) ────────────────────────────────
items_csv = '/home/diego/sources/BluenetERP/erpnext_items_import.csv'
item_fields = ['item_code', 'item_name', 'item_group', 'stock_uom', 'description']
with open(items_csv, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=item_fields, extrasaction='ignore')
    w.writeheader()
    for it in items:
        w.writerow(it)
print(f'{items_csv}  —  {len(items)} items')

# ── 3. Stock Reconciliation (opening stock) ────────────────────
stock_csv = '/home/diego/sources/BluenetERP/erpnext_stock_reconciliation_import.csv'
stock_fields = ['items.item_code', 'items.warehouse', 'items.qty']
with open(stock_csv, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=stock_fields)
    w.writeheader()
    for it in items:
        if it.get('opening_qty'):
            w.writerow({
                'items.item_code': it['item_code'],
                'items.warehouse': 'Guatemala',
                'items.qty': it['opening_qty'],
            })
print(f'{stock_csv}  —  {len(items)} stock entries (warehouse: Guatemala)')
