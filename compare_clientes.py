import openpyxl

wb = openpyxl.load_workbook("/home/diego/sources/BluenetERP/import/Clientes.xlsx")
ws = wb.active

header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

try:
    idx_nombre = header.index("nombre")
    idx_comercial = header.index("nombre_comercial")
except ValueError as e:
    print(f"Column not found: {e}")
    print(f"Available columns: {header}")
    exit(1)

diffs = 0
total = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    total += 1
    v1 = str(row[idx_nombre] or "").strip()
    v2 = str(row[idx_comercial] or "").strip()
    if v1 != v2:
        diffs += 1
        print(f"DIFF row {total+1}: nombre='{v1}' | nombre_comercial='{v2}'")

if diffs == 0:
    print(f"All {total} rows match: nombre and nombre_comercial are identical.")
else:
    print(f"\n{diffs} of {total} rows have differences.")
